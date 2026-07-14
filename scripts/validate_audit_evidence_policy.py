from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path


REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "backend"))

from app.services.audit_evidence_policy import (  # noqa: E402
    AUDIT_EVIDENCE_POLICY,
    RECOMMENDATION_EVIDENCE_DIMENSIONS,
    validate_audit_evidence_policy,
)


CAMPAIGN_TYPE_TO_SUBTYPE = {
    "search": "search",
    "brand_search": "brand_search",
    "yan": "yan_prospecting",
    "retargeting": "yan_retargeting",
    "master_campaign": "unknown",
    "mixed": "unknown",
}
ALLOWED_DESIGN_SPLIT = "synthetic_dev"
REGRESSION_SPLIT = "synthetic_test"
HOLDOUT_SPLIT = "synthetic_holdout"


def _codes(value: object) -> set[str]:
    return {
        item.strip()
        for item in re.split(r"[,;|\n]+", str(value or ""))
        if item.strip()
    }


def _headers(sheet) -> dict[str, int]:
    return {
        str(cell.value): index
        for index, cell in enumerate(sheet[1], start=1)
        if cell.value
    }


def _validate_case(
    *,
    row_number: int,
    split: str,
    case_type: str,
    campaign_type: str,
    recommendation_codes: set[str],
) -> list[str]:
    errors: list[str] = []
    subtype = CAMPAIGN_TYPE_TO_SUBTYPE.get(campaign_type)
    if subtype is None:
        errors.append(f"row {row_number} {split}: unknown campaign_type={campaign_type}")
        return errors
    if subtype not in AUDIT_EVIDENCE_POLICY.get(case_type, {}):
        errors.append(f"row {row_number} {split}: no policy for {case_type}+{subtype}")
    unknown_codes = recommendation_codes - set(RECOMMENDATION_EVIDENCE_DIMENSIONS)
    if unknown_codes:
        errors.append(f"row {row_number} {split}: unmapped recommendation codes={sorted(unknown_codes)}")
    return errors


def validate_dataset(path: Path) -> dict[str, int]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - depends on optional dev tooling
        raise SystemExit("Install openpyxl to validate the external XLSX dataset.") from exc

    workbook = load_workbook(path, read_only=False, data_only=True)
    required_sheets = {"Cases", "Code Dictionary"}
    missing_sheets = required_sheets - set(workbook.sheetnames)
    if missing_sheets:
        raise SystemExit(f"Missing workbook sheets: {sorted(missing_sheets)}")

    cases = workbook["Cases"]
    headers = _headers(cases)
    required_columns = {"dataset_split", "case_type", "campaign_type", "recommendation_codes"}
    missing_columns = required_columns - set(headers)
    if missing_columns:
        raise SystemExit(f"Missing Cases columns: {sorted(missing_columns)}")

    errors = validate_audit_evidence_policy()
    counts = {ALLOWED_DESIGN_SPLIT: 0, REGRESSION_SPLIT: 0, HOLDOUT_SPLIT: 0}
    for row_number in range(2, cases.max_row + 1):
        split = str(cases.cell(row_number, headers["dataset_split"]).value or "").strip()
        if not split:
            continue
        if split == HOLDOUT_SPLIT:
            # Deliberately do not access any other cell in holdout rows.
            counts[HOLDOUT_SPLIT] += 1
            continue
        if split not in {ALLOWED_DESIGN_SPLIT, REGRESSION_SPLIT}:
            errors.append(f"row {row_number}: unexpected dataset_split={split}")
            continue
        counts[split] += 1
        case_type = str(cases.cell(row_number, headers["case_type"]).value or "").strip()
        campaign_type = str(cases.cell(row_number, headers["campaign_type"]).value or "").strip()
        recommendation_codes = _codes(cases.cell(row_number, headers["recommendation_codes"]).value)
        errors.extend(_validate_case(
            row_number=row_number,
            split=split,
            case_type=case_type,
            campaign_type=campaign_type,
            recommendation_codes=recommendation_codes,
        ))

    dictionary = workbook["Code Dictionary"]
    dictionary_headers = _headers(dictionary)
    if not {"code", "category"} <= set(dictionary_headers):
        errors.append("Code Dictionary has no code/category columns")
    else:
        dictionary_codes = {
            str(dictionary.cell(row_number, dictionary_headers["code"]).value or "").strip()
            for row_number in range(2, dictionary.max_row + 1)
            if str(dictionary.cell(row_number, dictionary_headers["category"]).value or "").strip() == "recommendation"
        }
        dictionary_codes.discard("")
        missing_mappings = dictionary_codes - set(RECOMMENDATION_EVIDENCE_DIMENSIONS)
        if missing_mappings:
            errors.append(f"Code Dictionary contains unmapped codes: {sorted(missing_mappings)}")

    if errors:
        raise SystemExit("Audit evidence policy validation failed:\n- " + "\n- ".join(errors))
    return counts


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Validate deterministic audit evidence policy against an explicit external XLSX dataset.",
    )
    parser.add_argument("dataset_path", nargs="?", type=Path, help="Path to AI_Direct_Eval_Cases...xlsx")
    parser.add_argument("--dataset", dest="dataset_option", type=Path, help="Explicit external XLSX path")
    args = parser.parse_args()
    dataset = args.dataset_option or args.dataset_path
    if dataset is None or dataset.suffix.lower() != ".xlsx" or not dataset.is_file():
        raise SystemExit("Provide an existing XLSX dataset path.")
    counts = validate_dataset(dataset.resolve())
    print(
        "Audit evidence policy validated: "
        f"dev={counts[ALLOWED_DESIGN_SPLIT]}, "
        f"test={counts[REGRESSION_SPLIT]}, "
        f"holdout_count_only={counts[HOLDOUT_SPLIT]}."
    )


if __name__ == "__main__":
    main()
